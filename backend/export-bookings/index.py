import json
import os
import psycopg2
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import base64

def handler(event: dict, context) -> dict:
    '''Экспорт заявок в Excel файл с возможностью фильтрации по датам и статусу'''
    
    method = event.get('httpMethod', 'GET')
    
    if method == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'POST, OPTIONS',
                'Access-Control-Allow-Headers': 'Content-Type'
            },
            'body': ''
        }
    
    if method != 'POST':
        return {
            'statusCode': 405,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*'
            },
            'body': json.dumps({'error': 'Method not allowed'})
        }
    
    try:
        body = json.loads(event.get('body', '{}'))
        start_date = body.get('start_date')
        end_date = body.get('end_date')
        status_filter = body.get('status', 'all')
        
        database_url = os.environ.get('DATABASE_URL')
        schema_name = os.environ.get('MAIN_DB_SCHEMA')
        
        if not database_url or not schema_name:
            return {
                'statusCode': 500,
                'headers': {
                    'Content-Type': 'application/json',
                    'Access-Control-Allow-Origin': '*'
                },
                'body': json.dumps({'error': 'Ошибка конфигурации базы данных'})
            }
        
        conn = psycopg2.connect(database_url)
        cursor = conn.cursor()
        
        # Формируем запрос с фильтрами
        query_parts = [f"SELECT * FROM {schema_name}.bookings WHERE 1=1"]
        params = []
        
        if start_date and end_date:
            query_parts.append("AND created_at::date BETWEEN %s AND %s")
            params.extend([start_date, end_date])
        
        if status_filter != 'all':
            query_parts.append("AND status = %s")
            params.append(status_filter)
        
        query_parts.append("ORDER BY created_at DESC")
        query = " ".join(query_parts)
        
        cursor.execute(query, params)
        bookings = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        
        cursor.close()
        conn.close()
        
        # Создание Excel файла
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Заявки"
        
        # Русские названия колонок
        column_names = {
            'id': 'ID',
            'customer_name': 'Имя клиента',
            'customer_phone': 'Телефон',
            'customer_email': 'Email',
            'service_type': 'Тип услуги',
            'car_brand': 'Марка авто',
            'car_model': 'Модель авто',
            'preferred_date': 'Дата',
            'preferred_time': 'Время',
            'comment': 'Комментарий',
            'status': 'Статус',
            'created_at': 'Создано',
            'updated_at': 'Обновлено'
        }
        
        status_labels = {
            'new': 'Новая',
            'confirmed': 'Подтверждена',
            'completed': 'Завершена',
            'cancelled': 'Отменена'
        }
        
        # Заголовки
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = column_names.get(col_name, col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Данные
        for row_idx, booking in enumerate(bookings, 2):
            for col_idx, value in enumerate(booking, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Форматирование значений
                if columns[col_idx - 1] == 'status' and value:
                    cell.value = status_labels.get(value, value)
                elif isinstance(value, datetime):
                    cell.value = value.strftime('%d.%m.%Y %H:%M')
                elif value is None:
                    cell.value = '—'
                else:
                    cell.value = str(value)
                
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохранение в BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Кодирование в base64
        excel_base64 = base64.b64encode(excel_file.read()).decode('utf-8')
        
        # Формирование имени файла
        date_suffix = ''
        if start_date and end_date:
            date_suffix = f"_{start_date}_to_{end_date}"
        elif status_filter != 'all':
            date_suffix = f"_{status_filter}"
        
        filename = f"bookings{date_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return {
            'statusCode': 200,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*'
            },
            'body': json.dumps({
                'success': True,
                'file': excel_base64,
                'filename': filename,
                'count': len(bookings)
            })
        }
        
    except json.JSONDecodeError:
        return {
            'statusCode': 400,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*'
            },
            'body': json.dumps({'error': 'Неверный формат JSON'})
        }
    except Exception as e:
        return {
            'statusCode': 500,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*'
            },
            'body': json.dumps({'error': f'Ошибка сервера: {str(e)}'})
        }
